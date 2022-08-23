import openpyxl
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))

print(wb.sheetnames)
ws1 = wb.active
print(ws1['A3'].value)
print(ws1.cell(row=3, column=2).value)
ws1['A13'] = 'test2'
print(ws1['A13'].value)
wb.active.title = 'title'
wb.save('example.xlsx')